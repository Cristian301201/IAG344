import re
from openpyxl import load_workbook
# ===============================================
# Funcion clear_id
# Funcion para eliminar numeros y simbolos
# cc75.888.56 = "7588856"
# ===============================================

def clear_id ( value ):
    # Elimina caracteres no numericos
    if value is None:
        return "Se comio un payaso rey"
    return re.sub ( r"\D" , "" , str (value) ) #str() tranforma a string

# ================================================
# Funcion merge_name
# Une nombre y apellido en un solo campo
# ================================================

def merge_name (name, lastname) :
    if name is None:
        name = ""
    if lastname is None:
        lastname = ""
    return f"{name} {lastname}".strip() #Strip quita espacios 

# ================================================
# Funcion procces_excel
# ================================================

def process_excel (path) :
    # Acceso a la hoja llamada "datos"
    wb = load_workbook(path)
    ws = wb ["Datos"]
    # Recorrer todas las filas deade la numero 2
    for row in range (2 , ws.max_row + 1) :
        # Columna D: Identificador limpio
        ws [ f"D{row}" ] = clear_id ( ws [ f"A{row}"].value )
        # Colymna  E:  omnre completo
        ws [ f"E{row}" ] = merge_name ( 
        ws [ f"B{row}"].value , 
        ws [ f"C{row}"].value
        )
    #Guardar cambios archivo
    wb.save(path)

# ================================================
# Funcion process_excel_safe
# ================================================

def process_excel_safe ( path ) :
    try :
        process_excel ( path )
        return True , "Archivo procesado correctamente"
    except PermissionError : # Error cuando se intenta ejecutar con el archivo abierto
        return (
        False, "El archivo Excel está abierto.\n" # \n salta a la siguiente linea 
        "Por favor, cierre Excel e intene nuveamente." 
        )
    
    except KeyError : # Error cuando no encuentra la ruta a la hoja excel
        return False , "Hoja 'Datos' no encontrada"
    
    except Exception as e : # Errores no conocidos con variable temporal e
        return False , f"Error inesperado: {str(e)}"
